"""運送収支表Excel → mock/data.js 抽出スクリプト。

月次12シート・2024/2025年間集計・5月入力確認を読み取り、
モックが参照する window.APP_DATA を生成する。
"""
import datetime
import json
import re
import sys
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
XLSX = BASE / 'data' / '運送収支表 2025-2026 5月更新.xlsx'
OUT = BASE / 'mock' / 'data.js'

MONTH_SHEETS = [
    ('6月収支表', '2025-06'), ('7月収支表', '2025-07'), ('8月収支表', '2025-08'),
    ('9月収支表', '2025-09'), ('10月収支表', '2025-10'), ('11月収支表', '2025-11'),
    ('12月収支表', '2025-12'), ('1月収益表', '2026-01'), ('２月収支表', '2026-02'),
    ('3月収支表', '2026-03'), ('4月収支表', '2026-04'), ('5月収支表', '2026-05'),
]

# 51列の正準キー(列順どおり)
FIELDS = [
    'no', 'type', 'depot', 'reg', 'code', 'driver',
    'trips', 'slips', 'hours', 'km',
    'fare', 'fee', 'sales',
    'toll', 'tollDisc', 'tollNet',
    'fuelIn', 'fuelInQty', 'fuelOut', 'fuelOutQty', 'fuelQty', 'nempi', 'adblue', 'fuelTotal',
    'repair', 'tire', 'equip', 'mainte', 'repairTotal',
    'salary', 'bonus', 'welfare', 'laborTotal',
    'insCompulsory', 'insVoluntary', 'insTotal',
    'taxAuto', 'taxWeight', 'taxTotal',
    'miscOther', 'miscTotal',
    'lease', 'installment', 'transportTotal',
    'adminFee', 'adminTotal',
    'fixed', 'variable', 'expense', 'profit', 'margin',
]
NUM_START = 6  # trips以降は数値列


def norm(s):
    if s is None:
        return ''
    return re.sub(r'[\s　]+', '', str(s))


def clean(v, idx):
    """セル値をJSON化可能な値に正規化する。"""
    if v is None:
        return None
    if isinstance(v, datetime.timedelta):
        return round(v.total_seconds() / 3600, 1)  # 稼働時間 → 時間
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m')
    if isinstance(v, str):
        s = v.strip()
        if s in ('', '#DIV/0!', '#REF!', '#VALUE!', '#N/A'):
            return None
        if idx >= NUM_START:
            try:
                return float(s.replace(',', ''))
            except ValueError:
                return s
        return re.sub(r'[　]+', ' ', s).strip()
    if isinstance(v, float):
        return round(v, 4)
    return v


def extract_month(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = [norm(c) for c in rows[2][:51]]
    if header[0] != '車番' or not header[49].startswith('損益'):
        raise SystemExit(f'{ws.title}: 想定外ヘッダー {header[:3]} … {header[49]}')
    tank_price = None
    for c in rows[0]:
        if isinstance(c, str) and 'インタンク' in c:
            m = re.search(r'([\d.]+)円', c)
            if m:
                tank_price = float(m.group(1))
    vehicles = []
    for r in rows[3:]:
        a = norm(r[0])
        if a == '合計':
            break
        if a == '' and norm(r[1]) == '' and norm(r[5]) == '':
            continue
        rec = [clean(r[i], i) for i in range(51)]
        rec[0] = str(rec[0]) if rec[0] is not None else ''
        vehicles.append(rec)
    return tank_price, vehicles


def extract_annual(ws):
    """年間集計シートの【全体】ブロック(6月〜5月)を抽出する。"""
    rows = list(ws.iter_rows(values_only=True))
    out = []
    keys = ['month', 'unko', 'fuel', 'repair', 'labor', 'ins', 'tax', 'transport',
            'admin', 'total', 'km', 'costPerKm', 'sales', 'salesPerKm', 'cars', 'days', 'slips']
    for r in rows[2:15]:
        label = norm(r[0])
        if not re.match(r'^\d+月$', label):
            continue
        rec = {}
        for i, k in enumerate(keys):
            v = r[i]
            if isinstance(v, str) and v.startswith('#'):
                v = None
            if isinstance(v, float):
                v = round(v, 2)
            rec[k] = label if i == 0 else v
        out.append(rec)
    return out


def extract_anomalies(ws):
    keys = ['priority', 'no', 'type', 'driver', 'cell', 'item', 'value', 'guide', 'reason']
    out = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not any(r):
            continue
        out.append({k: ('' if v is None else str(v).strip()) for k, v in zip(keys, r)})
    return out


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    months = []
    for sheet, ym in MONTH_SHEETS:
        tank, vehicles = extract_month(wb[sheet])
        months.append({
            'ym': ym,
            'label': f'{int(ym[5:])}月',
            'sheet': sheet,
            'tankPrice': tank,
            'rows': vehicles,
        })
        print(f'{sheet}: {len(vehicles)}台 tank={tank}')

    data = {
        'fields': FIELDS,
        'months': months,
        'annual2024': extract_annual(wb['2024']),
        'annual2025Sheet': extract_annual(wb['2025']),
        'anomalies': extract_anomalies(wb['5月入力確認']),
        'sourceFile': XLSX.name,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    js = 'window.APP_DATA = ' + json.dumps(data, ensure_ascii=False, separators=(',', ':')) + ';\n'
    OUT.write_text(js, encoding='utf-8')
    print(f'written: {OUT} ({OUT.stat().st_size / 1024:.0f} KB)')


if __name__ == '__main__':
    sys.exit(main())
